©Spencer Hiscox, 2023

import re
from datetime import datetime as dt
from openpyxl import load_workbook

#determine current date
class date_info:
    def __init__(self, day, month, year):
        self.year = year
        self.month = month
        self.day = day
        
    def disp(self):
        return str(self.month) + "/" + str(self.day) + "/" + str(self.year)

global current_date
current_date = date_info(dt.date(dt.now()).day, 
                         dt.date(dt.now()).month, 
                         dt.date(dt.now()).year)

class t2cc:
    def __init__(self, weeks, days):
        self.weeks = weeks
        self.days = days
        
    def disp(self):
        return f"weeks = {self.weeks}, days = {self.days}"
    
test_case = t2cc(3, 2)
test_case = type(test_case)

def calculate_weeks_until_due(obj, assigned_time=True):
    days_in_month = {1 : 31, 3 : 31, 4 : 30, 5 : 31, 6 : 30, 7 : 31,
                     8 : 31, 9 : 30, 10 : 31, 11 : 30, 12 : 31}
    if not current_date.year % 4:
        days_in_month[2] = 29
    else:
        days_in_month[2] = 28
    weeks, days = 0, 0
    if assigned_time:
        if obj.dd.month > obj.ad.month:
            weeks = (obj.dd.month - obj.ad.month - 1) * 4
            weeks += obj.dd.day // 7
            days = obj.dd.day % 7
            weeks += (days_in_month[obj.ad.month] - obj.ad.day) // 7
            days += (days_in_month[obj.ad.month] - obj.ad.day) % 7
        elif obj.dd.month == current_date.month:
            if obj.dd.day > obj.ad.day:
                weeks = (obj.dd.day - obj.ad.day) // 7
                days = (obj.dd.day - obj.ad.day) % 7
            else:
                weeks = (obj.ad.day - obj.dd.day) // 7
                days = -((obj.ad.day - obj.dd.day) % 7)
        else:
            weeks = 0
            days = 0
        return t2cc(weeks, days)
    else:
        if obj.dd.month > current_date.month:
            weeks = (obj.dd.month - current_date.month - 1) * 4
            weeks += obj.dd.day // 7
            days = obj.dd.day % 7
            weeks += (days_in_month[current_date.month] - current_date.day) // 7
            days += (days_in_month[current_date.month] - current_date.day) % 7
        elif obj.dd.month == current_date.month:
            if obj.dd.day > obj.ad.day:
                weeks = (obj.dd.day - current_date.day) // 7
                days = (obj.dd.day - current_date.day) % 7
            else:
                weeks = (current_date.day - obj.dd.day) // 7
                days = -((current_date.day - obj.dd.day) % 7)
        else:
            weeks = 0
            days = 0
        return t2cc(weeks, days)        

class task:
    def __init__(self, auth, subject, title, typ, assdate, duedate, duetime, hrs, complete):
        self.auth = auth        #COURSE / COOP / SDT / ANCILLARY
        self.sub = subject      #WHICH SDT or WHICH COURSE or WHICH COOP (coop root / CFE etc.)
        self.name = title       #name of task
        self.typ = typ          #type of task (reading / assignment (/todo item) / test or quiz or exam etc.
        self.ad = assdate       #if we run this everyday, then the day it finds the thing to import can be taken as the assigned date
        self.dd = duedate       #due date
        self.dt = duetime       #due-by time (in the excel doc, should always reflect submission deadline (not my bedtime))
        self.hrs = hrs          #estimated amount of time it will take to complete the task
        self.done = complete    #just a boolean that marks a task as complete or not complete **CAN ONLY ONLY BE UPDATED WITHIN THIS VIEWER
        if self.dd:
            self.t2c = calculate_weeks_until_due(self, True)
            self.tr = calculate_weeks_until_due(self, False)

global current_tasks, completed_tasks
current_tasks, completed_tasks = [], []

global auth, acad_subject, coop_subject, sdt_subject, type_list, subject_list
auth = ("COURSE", "COOP", "SDT", "ANCILLARY")
acad_subject = ("MATH 116 (Calculus)", "MATH 115 (Linear Algebra)", "PHYS 115", "CHE 102", "GENE 119", "ME 100")
coop_subject = ("CFE", "COOP", "Engineering Co-op Community", "other")
sdt_subject = ("UW Orbital", "Formula SAE")
type_list = {'A' : 'ASSIGNMENT', 'E' : 'EXAM', 'R' : 'READING'}

subject_list = {auth[0] : acad_subject, auth[1] : coop_subject, auth[2] : sdt_subject, auth[3] : []}

#start by checking if there's a csv in the data file we can import from? -- then, import to data structure, following saaammme-ish protocol as importing from deadlines.xlsx
#then we can just import *new* items from deadlines.xlxs
#should, therefore, define a function to determine if an entry is the same as an existing entry -- all this allows date assigned to be automated -- though I should write in the ability
#to manually enter the date assigned

#let's start with importing and displaying -- possible putting in a REALLY basic menu that can change the display organization or change the assigned date on something -- which should then update
#the display

def read_excel():
    global current_date, current_tasks, completed_tasks, auth, acad_subject, coop_subject, sdt_subject, type_list, subject_list
    
    wb = load_workbook(filename="C:\\Users\\shisc\\My Drive\\Personal\\2 - University of Waterloo\\Academics\\Undergrad\\Y1_F23\\Deadlines.xlsx")
    authority, auth_updated = "", False
    subject = ""
    duplicate = False
    
    raw_data, sheet = [], wb.worksheets[0]
    max_row = int(re.findall('[\d]*(?=$)', sheet.calculate_dimension())[0]) + 1
    
    for i in range(1, max_row):
        col_values = []
        for j in range(1, len(sheet[i]) + 1):
            if j != len(sheet[i]):
                if sheet.cell(i, j).value or sheet.cell(i, j + 1).value:
                    col_values += [sheet.cell(i, j).value]
            else:
                if sheet.cell(i, j).value:
                    col_values += [sheet.cell(i, j).value]
        if col_values:
            raw_data += [col_values]
    
    for data in raw_data:
        if data[0]:
            for key in auth:                                                #figure out which auth we're dealing with
                if re.findall(key, data[0]):
                    authority = key
                    auth_updated = True
                    break
            if auth_updated:
                auth_updated = False
                continue
            for sub in subject_list[authority]:                                  #figure out which "subject" under that auth we're dealing with
                search_term = ""
                if re.findall("\(", sub):
                    end_index = sub.index("(") - 1
                    search_term = sub[:end_index]
                else:
                    search_term = sub
                if re.findall(search_term, data[0]):
                    subject = sub
                    break
        if len(data) > 6:
            date_due = (data[2].month, data[2].day, data[2].year)
            temp = task(authority, 
                        subject, 
                        data[1],
                        type_list[data[5]], 
                        date_info(current_date.day, 
                                  current_date.month, 
                                  current_date.year), 
                        date_info(date_due[1], 
                                  date_due[0], 
                                  date_due[2]), 
                        data[3], 
                        data[4], 
                        data[6] == 'X')
        
        else:
            if (data[0] not in auth) and (raw_data.index(data) < 45):
                temp = task(authority, subject, None, None, None, None, None, None, False)
            
        if temp.done:
            if completed_tasks:
                for item in completed_tasks:
                    if item == temp:
                        duplicate = True
                        break
                if duplicate:
                    duplicate = False
                    continue
            if not temp.name and \
               not temp.typ and \
               not temp.ad and \
               not temp.dd and \
               not temp.dt and \
               not temp.hrs and \
               not temp.done and \
               temp.auth == 'COURSE':           #THIS (and below) EXCLUSIONARY BIT (to get rid of empty lines) Is causing a problem with the SDT, COOP and ANCILLARY areas
                continue                        #I *think* we're still in the nested for loops here though, so we could look ahead to the next 'subject' (down through rows) potentially...
            completed_tasks += [temp]
        else:
            if current_tasks:
                for item in current_tasks:
                    if item == temp:
                        duplicate = True
                        break
                if duplicate:
                    duplicate = False
                    continue
            if not temp.name and \
               not temp.typ and \
               not temp.ad and \
               not temp.dd and \
               not temp.dt and \
               not temp.hrs and \
               not temp.done and\
               temp.auth == 'COURSE':
                continue            
            current_tasks += [temp]

def display_current_tasks(view=True):
    if view:
        #display by duedate
        #dont forget, splitting by how much time was originally given for the assignment, obj.t2c.weeks = 0 and obj.t2c.weeks = 1 should both come under 1 week (or could have a <1wk category)
        pass
    else:
        #display by course
        pass

read_excel()

type_check = type(current_date)
for entry in range(len(current_tasks)):                                         #Got a few data reading errors to fix in read_excel() before we can go any further.
    for key in current_tasks[entry].__dict__.keys():
        if type(current_tasks[entry].__dict__[key]) == type_check or \
           type(current_tasks[entry].__dict__[key]) == test_case:
            print(key, current_tasks[entry].__dict__[key].disp(), sep="\t")
        else:
            print(key, current_tasks[entry].__dict__[key], sep="\t")
    print()

for entry in range(len(completed_tasks)):                                         #Got a few data reading errors to fix in read_excel() before we can go any further.
    for key in completed_tasks[entry].__dict__.keys():
        if type(completed_tasks[entry].__dict__[key]) == type_check or \
           type(completed_tasks[entry].__dict__[key]) == test_case:
            print(key, completed_tasks[entry].__dict__[key].disp(), sep="\t")
        else:
            print(key, completed_tasks[entry].__dict__[key], sep="\t")
    print()
 
"""

    
    for row in range(1, max_row):
        col_values = []
        none_counter = 0
        for cell in sheet[row]:
            if cell.value == None:
                none_counter += 1
                if none_counter > 1:
                    break
            col_values += [cell.value]
        if col_values:
            raw_data += [col_values]
"""
